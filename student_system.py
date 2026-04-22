from openpyxl import load_workbook, Workbook
import os
from tkinter import *
from tkinter import messagebox

file = "students.xlsx"

if not os.path.exists(file):
    wb = Workbook()
    ws = wb.active
    headers = ["ID", "Name", "Course", "Phone"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = h
    wb.save(file)

def create_xl():
    wb_local = load_workbook(file)
    ws_local = wb_local.active
    headers = ["ID", "Name", "Course", "Phone"]

    for i, h in enumerate(headers, 1):
        if ws_local.cell(row=1, column=i).value != h:
            ws_local.cell(row=1, column=i).value = h
    wb_local.save(file)

create_xl()

def focus_next(s):
    s.focus_set()

def clear():
    for k in entries:
        k.delete(0, END)

def add_student():
    if all(t.get() for t in entries):
        wb_local = load_workbook(file)
        ws_local = wb_local.active
        row = ws_local.max_row + 1
        for i, t in enumerate(entries, 1):
            ws_local.cell(row=row, column=i).value = t.get()
        wb_local.save(file)
        clear()
        messagebox.showinfo("Done", "Student added successfully")
    else:
        messagebox.showwarning('Warning', "Please fill all items.")

# search
def search_std():
    std_id = search_id.get()
    if std_id == "":
        messagebox.showerror("Error", "Please enter the ID")
        return
    wb_local = load_workbook(file)
    sr = wb_local.active
    clear()
    found = False
    for row in sr.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(std_id):
            found = True
            for i, value in enumerate(row):
                entries[i].insert(0, value)
            break
    if not found:
        messagebox.showinfo("Not found", "Not found student")
    return

# delete
def delete_std():
    std_id = del_id.get()
    if std_id == "":
        messagebox.showerror("Error", "Please enter the ID")
        return
    wb_local = load_workbook(file)
    d = wb_local.active
    found = False
    for r in range(2, d.max_row + 1):
        if str(d.cell(row=r, column=1).value) == str(std_id):
            d.delete_rows(r)
            wb_local.save(file)
            found = True
            messagebox.showinfo("Deleted", "Successfully deleted")
            break
    if not found:
        messagebox.showinfo("Not found", "Not found student")

# update
def update_std():
    std_id = up_std.get()
    if std_id == "":
        messagebox.showerror("Error", "Please enter the ID")
        return
    wb_local = load_workbook(file)
    sr = wb_local.active
    found = False
    for r in range(2, sr.max_row + 1):
        if str(sr.cell(row=r, column=1).value) == str(std_id):
            for i, t in enumerate(entries, 1):
                sr.cell(row=r, column=i).value = t.get()
            wb_local.save(file)
            found = True
            messagebox.showinfo("Updated", "Successfully updated")
            clear()
            break
    if not found:
        messagebox.showinfo("Not found", "Not found student")

#Show All
def show_all():
    listbox.delete(0, END)
    wb_local = load_workbook(file)
    sh = wb_local.active
    for row in sh.iter_rows(min_row=2, values_only=True):
        display = f"{row[0]} | {row[1]} | {row[2]} | {row[3]}"
        listbox.insert(END, display)

#GUI
student = Tk()
student.title("Student Management System")
student.geometry("700x450+100+50")
student.minsize(700, 450)
student.config(bg="light blue")

Label(student, text="Registration", bg="light blue", font=("Arial", 14)).grid(row=0, column=1, pady=5)

std = ["ID", "Name", "Course", "Phone"]
entries = [Entry(student) for _ in std]

for i, lab in enumerate(std):
    Label(student, text=lab).grid(row=i + 1, column=0, sticky=W, padx=5, pady=3)
    entries[i].grid(row=i + 1, column=1, ipadx=40, pady=3)
    if i < len(std) - 1:
        entries[i].bind("<Return>", lambda a, nb=entries[i + 1]: focus_next(nb))


search_id = Entry(student)
search_id.grid(row=1, column=4, padx=5)
del_id = Entry(student)
del_id.grid(row=2, column=4, padx=5)
up_std = Entry(student)
up_std.grid(row=3, column=4, padx=5)

Button(student, text="Add Student", fg="black", command=add_student).grid(row=5, column=1, pady=5)
Button(student, text="Search", fg="black", command=search_std).grid(row=1, column=3, padx=5)
Button(student, text="Delete", fg="black", command=delete_std).grid(row=2, column=3, padx=5)
Button(student, text="Update", fg="black", command=update_std).grid(row=3, column=3, padx=5)


Button(student, text="Show All Students", command=show_all).grid(row=9, column=2, pady=4)
listbox = Listbox(student, width=110, height=10)
listbox.grid(row=8, column=0, columnspan=7, pady=30, padx=10)

student.mainloop()
